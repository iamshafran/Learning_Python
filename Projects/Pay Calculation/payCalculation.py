import os.path
import pickle
from datetime import datetime, time, timedelta
import holidays
from apiclient.discovery import build
from dateutil.parser import *
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import pprint

# get shifts form google calendar
SCOPES = ["https://www.googleapis.com/auth/calendar.readonly"]

creds = None
# The file token.pickle stores the user's access and refresh tokens, and is
# created automatically when the authorization flow completes for the first
# time.
if os.path.exists("token.pickle"):
    with open("token.pickle", "rb") as token:
        creds = pickle.load(token)
# If there are no (valid) credentials available, let the user log in.
if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    else:
        flow = InstalledAppFlow.from_client_secrets_file(
            "/Users/iamshafran/Development/Python/Projects/Pay Calculation/credentials.json",
            SCOPES,
        )
        creds = flow.run_local_server(port=0)
    # Save the credentials for the next run
    with open("token.pickle", "wb") as token:
        pickle.dump(creds, token)

service = build("calendar", "v3", credentials=creds)

payday = datetime.strptime(
    input("Please enter the payday in format YYYY-MM-DD:"), "%Y-%m-%d"
)
hourlyPay = 11.614
nightPremium = 2.211

endDate = (payday - timedelta(7)).isoformat() + "Z"
startDate = (payday - timedelta(35)).isoformat() + "Z"

result = (
    service.events()
    .list(
        calendarId="n35549an1lat7be1b9i17ajeuo@group.calendar.google.com",
        timeMin=startDate,
        timeMax=endDate,
        singleEvents=True,
        orderBy="startTime",
    )
    .execute()
)

events = result.get("items", [])

englandHoliday = holidays.England()

weeklyHours = {"week1": 0, "week2": 0, "week3": 0, "week4": 0}


# determine shift pattern and hours worked
for event in events:
    # Get datetime from Google Calendar data and strip timezone info
    # pprint.pprint(event)
    event["start"]["dateTime"] = parse(event["start"].get("dateTime")).replace(
        tzinfo=None
    )
    event["end"]["dateTime"] = parse(event["end"].get("dateTime")).replace(tzinfo=None)

    event["start"]["startTime"] = (event["start"].get("dateTime")).strftime("%H:%M:%S")
    event["end"]["endTime"] = (event["end"].get("dateTime")).strftime("%H:%M:%S")

    event["start"]["startDate"] = (event["start"].get("dateTime")).strftime("%Y-%m-%d")
    event["end"]["endDate"] = (event["end"].get("dateTime")).strftime("%Y-%m-%d")

    event["hoursWorked"] = (
        ((parse(event["end"]["endTime"]) - parse(event["start"]["startTime"])).seconds)
        / 3600
    ) - 0.5
    event["bankHoliday"] = event["start"].get("dateTime") in englandHoliday

    if parse(event["start"]["startTime"]).time() < time(23, 59) and parse(
        event["end"]["endTime"]
    ).time() < time(10, 0):
        event["shift"] = "night"
        if parse(event["end"]["endTime"]).time() < time(6, 1):
            event["nightHours"] = (
                (parse(event["end"]["endTime"]) - parse("00:00:00")).seconds
            ) / 3600

    elif parse(event["start"]["startTime"]).time() > time(5, 59) and parse(
        event["end"]["endTime"]
    ).time() < time(19, 1):
        event["shift"] = "early"

    elif parse(event["start"]["startTime"]).time() > time(10, 59) and parse(
        event["end"]["endTime"]
    ).time() < time(22, 59):
        event["shift"] = "late"

# determine week numbers. Week starting on Friday and ending on Thursdays
for event in events:

    eventDate = event["start"].get("dateTime")
    dateDifference = (payday - eventDate.replace(tzinfo=None)).days

    if dateDifference <= 35 and dateDifference >= 29:
        event["weekNumber"] = 1

    elif dateDifference <= 28 and dateDifference >= 22:
        event["weekNumber"] = 2

    elif dateDifference <= 21 and dateDifference >= 15:
        event["weekNumber"] = 3

    elif dateDifference <= 14 and dateDifference >= 8:
        event["weekNumber"] = 4

# calculate weekly hours worked
for i in range(1, 5):
    shiftCount = 1
    for event in events:
        if event["weekNumber"] == i:
            if shiftCount > 5:
                event["otShift"] = True
            else:
                event["otShift"] = False
            shiftCount += 1
            weeklyHours["week" + str(i)] = (
                weeklyHours["week" + str(i)] + event["hoursWorked"]
            )

    # check for minus hours
    if weeklyHours["week" + str(i)] < 36.5:
        event["weeklyMinusHours"] = weeklyHours["week" + str(i)] - 36.5

# calculate holiday pay
for week in range(1, 5):
    contractHours = 36.5
    for event in events:
        if week == event["weekNumber"]:
            if event["summary"] == "Holidays":
                if contractHours > 6.5:
                    event["holidayPay"] = hourlyPay * 7.5
                else:
                    event["holidayPay"] = hourlyPay * contractHours
            contractHours = contractHours - 7.5

# calculate sick pay
for week in range(1, 5):
    contractHours = 36.5
    sickCounter = 1
    for event in events:
        if week == event["weekNumber"]:
            if event["summary"] == "Sick":
                if sickCounter < 4:
                    if contractHours > 6.5:
                        event["minusHours"] = 7.5
                    else:
                        event["minusHours"] = contractHours
            contractHours = contractHours - 7.5

# calculate pay
for event in events:
    # determine overtime and calculate daily pay
    if event["otShift"]:
        event["otPay"] = event["hoursWorked"] * hourlyPay

    elif event["hoursWorked"] > 7.5:
        event["otHours"] = event["hoursWorked"] - 7.5
        event["otPay"] = event["otHours"] * hourlyPay
        event["dailyPay"] = 7.5 * hourlyPay

    else:
        event["dailyPay"] = event["hoursWorked"] * hourlyPay

    try:
        if event.get("minusHours"):
            event["dailyPay"] = 0
    except KeyError:
        pass

    # calculate bank holiday pay
    if event["bankHoliday"]:
        event["dailyPay"] = event["dailyPay"] * 1.25

    # calculate Sunday premium
    if event["start"].get("dateTime").weekday() == 6:
        event["sunday"] = True
        event["premiums"] = {"sundayPremium": event["dailyPay"] * 0.25}

    # calculate night premium
    if event.get("nightHours"):
        event["premiums"] = {"nightPremium": event["nightHours"] * nightPremium}

shiftSort = {"week1": {}, "week2": {}, "week3": {}, "week4": {}}

for week in range(1, 5):
    for event in events:
        if event["weekNumber"] == week:
            shiftSort["week" + str(week)] = {"day": event["start"].get("dateTime")}

wb = load_workbook("/Users/iamshafran/Pay Calculations/Template.xlsx")
ws = wb.active

row = 7
# write to excel file
for event in events:
    ws["A" + str(row)] = event["start"].get("dateTime")
    ws["B" + str(row)] = str(event.get("shift")).capitalize()
    if event["bankHoliday"]:
        ws["C" + str(row)] = "Y"
    else:
        ws["C" + str(row)] = "N"
    ws["D" + str(row)] = event["start"].get("startTime")
    ws["E" + str(row)] = event["end"].get("endTime")
    ws["F" + str(row)] = event.get("hoursWorked")
    if event["bankHoliday"]:
        ws["K" + str(row)] = event.get("dailyPay")
    else:
        ws["G" + str(row)] = event.get("dailyPay")
    ws["H" + str(row)] = event.get("otPay")
    try:
        ws["I" + str(row)] = event["premiums"].get("sundayPremium")
        ws["J" + str(row)] = event["premiums"].get("nightPremium")
    except KeyError:
        pass
    ws["L" + str(row)] = event.get("holidayPay")
    ws["M" + str(row)] = event.get("minusHours")
    if shiftSort["week" + str(event["weekNumber"])].get("day") == event["start"].get(
        "dateTime"
    ):
        ws["M" + str(row)] = event.get("weeklyMinusHours")
    ws["N" + str(row)] = "=SUM(G" + str(row) + ":L" + str(row) + ")"

    if shiftSort["week" + str(event["weekNumber"])].get("day") == event["start"].get(
        "dateTime"
    ):
        for column in range(65, 79):
            ws[chr(column) + str(row)].fill = PatternFill(
                start_color="FEFFA6", end_color="FEFFA6", fill_type="solid"
            )

    row += 1

ws["C4"] = hourlyPay
ws["G35"] = "=SUM(G7:G34)"
ws["H35"] = "=SUM(H7:H34)"
ws["I35"] = "=SUM(I7:I34)"
ws["J35"] = "=SUM(J7:J34)"
ws["K35"] = "=SUM(K7:K34)"
ws["L35"] = "=SUM(L7:L34)"
ws["M35"] = "=SUM(M7:M34)"
ws["N35"] = "=SUM(N7:N34)"

wb.save("/Users/iamshafran/Pay Calculations/" + payday.strftime("%B") + ".xlsx")